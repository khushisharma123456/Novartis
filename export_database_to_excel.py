"""
Export complete database to Excel with multiple sheets
Creates a well-organized Excel workbook with all database tables
"""

import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app import app, db
from models import User, Patient, Drug, Alert, SideEffectReport, hospital_doctor, hospital_drug, hospital_pharmacy, doctor_patient

def style_header(worksheet):
    """Apply styling to header row"""
    header_fill = PatternFill(start_color="1F3A52", end_color="1F3A52", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

def auto_adjust_columns(worksheet):
    """Auto-adjust column widths"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Max width of 50
        worksheet.column_dimensions[column_letter].width = adjusted_width

def export_users(wb):
    """Export Users table"""
    ws = wb.create_sheet("Users")
    
    # Headers
    headers = ["ID", "Name", "Email", "Role", "Hospital Name"]
    ws.append(headers)
    style_header(ws)
    
    # Data
    users = User.query.all()
    for user in users:
        ws.append([
            user.id,
            user.name,
            user.email,
            user.role,
            user.hospital_name or ""
        ])
    
    auto_adjust_columns(ws)
    return len(users)

def export_drugs(wb):
    """Export Drugs table"""
    ws = wb.create_sheet("Drugs")
    
    # Headers
    headers = ["ID", "Name", "Company", "Description", "Active Ingredients", 
               "AI Risk Assessment", "AI Risk Details", "Created At"]
    ws.append(headers)
    style_header(ws)
    
    # Data
    drugs = Drug.query.all()
    for drug in drugs:
        ws.append([
            drug.id,
            drug.name,
            drug.company.name if drug.company else "",
            drug.description or "",
            drug.active_ingredients or "",
            drug.ai_risk_assessment or "",
            drug.ai_risk_details or "",
            drug.created_at.strftime("%Y-%m-%d %H:%M:%S") if drug.created_at else ""
        ])
    
    auto_adjust_columns(ws)
    return len(drugs)

def export_patients(wb):
    """Export Patients table"""
    ws = wb.create_sheet("Patients")
    
    # Headers
    headers = ["ID", "Name", "Age", "Gender", "Phone", "Drug Name", "Symptoms", "Risk Level", 
               "Case Status", "Match Score", "Recalled", "Recalled By", "Recall Reason", 
               "Recall Date", "Created By", "Created At"]
    ws.append(headers)
    style_header(ws)
    
    # Data
    patients = Patient.query.all()
    for patient in patients:
        recalled_by_user = User.query.get(patient.recalled_by) if patient.recalled_by else None
        created_by_user = User.query.get(patient.created_by) if patient.created_by else None
        
        ws.append([
            patient.id,
            patient.name,
            patient.age,
            patient.gender,
            patient.phone or "",
            patient.drug_name or "",
            patient.symptoms or "",
            patient.risk_level or "",
            patient.case_status or "",
            patient.match_score or "",
            "Yes" if patient.recalled else "No",
            recalled_by_user.name if recalled_by_user else "",
            patient.recall_reason or "",
            patient.recall_date.strftime("%Y-%m-%d %H:%M:%S") if patient.recall_date else "",
            created_by_user.name if created_by_user else "",
            patient.created_at.strftime("%Y-%m-%d %H:%M:%S") if patient.created_at else ""
        ])
    
    auto_adjust_columns(ws)
    return len(patients)

def export_alerts(wb):
    """Export Alerts table"""
    ws = wb.create_sheet("Alerts")
    
    # Headers
    headers = ["ID", "Drug Name", "Title", "Sender", "Message", "Severity", "Recipient Type", 
               "Is Read", "Created At"]
    ws.append(headers)
    style_header(ws)
    
    # Data
    alerts = Alert.query.all()
    for alert in alerts:
        sender = User.query.get(alert.sender_id) if alert.sender_id else None
        
        ws.append([
            alert.id,
            alert.drug_name or "",
            alert.title or "",
            sender.name if sender else "",
            alert.message or "",
            alert.severity or "",
            alert.recipient_type or "",
            "Yes" if alert.is_read else "No",
            alert.created_at.strftime("%Y-%m-%d %H:%M:%S") if alert.created_at else ""
        ])
    
    auto_adjust_columns(ws)
    return len(alerts)

def export_relationships(wb):
    """Export relationship tables"""
    
    # Hospital-Doctor relationships
    ws_hd = wb.create_sheet("Hospital-Doctor Links")
    ws_hd.append(["Hospital ID", "Hospital Name", "Doctor ID", "Doctor Name"])
    style_header(ws_hd)
    
    with app.app_context():
        results = db.session.execute(db.select(hospital_doctor)).fetchall()
        hd_count = 0
        for row in results:
            hospital = User.query.get(row.hospital_id)
            doctor = User.query.get(row.doctor_id)
            if hospital and doctor:
                ws_hd.append([
                    hospital.id,
                    hospital.name,
                    doctor.id,
                    doctor.name
                ])
                hd_count += 1
    auto_adjust_columns(ws_hd)
    
    # Hospital-Drug relationships
    ws_drug = wb.create_sheet("Hospital-Drug Links")
    ws_drug.append(["Hospital ID", "Hospital Name", "Drug ID", "Drug Name", "Company"])
    style_header(ws_drug)
    
    with app.app_context():
        results = db.session.execute(db.select(hospital_drug)).fetchall()
        drug_count = 0
        for row in results:
            hospital = User.query.get(row.hospital_id)
            drug = Drug.query.get(row.drug_id)
            if hospital and drug:
                ws_drug.append([
                    hospital.id,
                    hospital.name,
                    drug.id,
                    drug.name,
                    drug.company.name if drug.company else ""
                ])
                drug_count += 1
    auto_adjust_columns(ws_drug)
    
    # Hospital-Pharmacy relationships
    ws_pharm = wb.create_sheet("Hospital-Pharmacy Links")
    ws_pharm.append(["Hospital ID", "Hospital Name", "Pharmacy ID", "Pharmacy Name"])
    style_header(ws_pharm)
    
    with app.app_context():
        results = db.session.execute(db.select(hospital_pharmacy)).fetchall()
        pharm_count = 0
        for row in results:
            hospital = User.query.get(row.hospital_id)
            pharmacy = User.query.get(row.pharmacy_id)
            if hospital and pharmacy:
                ws_pharm.append([
                    hospital.id,
                    hospital.name,
                    pharmacy.id,
                    pharmacy.name
                ])
                pharm_count += 1
    auto_adjust_columns(ws_pharm)
    
    # Doctor-Patient relationships
    ws_dp = wb.create_sheet("Doctor-Patient Links")
    ws_dp.append(["Doctor ID", "Doctor Name", "Patient ID", "Patient Name", "Drug Name", "Risk Level"])
    style_header(ws_dp)
    
    with app.app_context():
        results = db.session.execute(db.select(doctor_patient)).fetchall()
        dp_count = 0
        for row in results:
            doctor = User.query.get(row.doctor_id)
            patient = Patient.query.get(row.patient_id)
            if doctor and patient:
                ws_dp.append([
                    doctor.id,
                    doctor.name,
                    patient.id,
                    patient.name,
                    patient.drug_name or "",
                    patient.risk_level or ""
                ])
                dp_count += 1
    auto_adjust_columns(ws_dp)
    
    return hd_count, drug_count, pharm_count, dp_count

def export_summary(wb, stats):
    """Create summary sheet"""
    ws = wb.create_sheet("Summary", 0)  # Insert as first sheet
    
    # Title
    ws.merge_cells('A1:B1')
    ws['A1'] = "InteLeYzer Database Export"
    ws['A1'].font = Font(size=16, bold=True, color="1F3A52")
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:B2')
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # Stats
    ws['A4'] = "Table"
    ws['B4'] = "Record Count"
    ws['A4'].font = Font(bold=True)
    ws['B4'].font = Font(bold=True)
    
    row = 5
    for table, count in stats.items():
        ws[f'A{row}'] = table
        ws[f'B{row}'] = count
        row += 1
    
    # Auto-adjust
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15

def main():
    print("\n" + "="*60)
    print("EXPORTING DATABASE TO EXCEL")
    print("="*60 + "\n")
    
    with app.app_context():
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Export all tables
        print("📊 Exporting Users...")
        user_count = export_users(wb)
        print(f"   ✓ Exported {user_count} users")
        
        print("📊 Exporting Drugs...")
        drug_count = export_drugs(wb)
        print(f"   ✓ Exported {drug_count} drugs")
        
        print("📊 Exporting Patients...")
        patient_count = export_patients(wb)
        print(f"   ✓ Exported {patient_count} patients")
        
        print("📊 Exporting Alerts...")
        alert_count = export_alerts(wb)
        print(f"   ✓ Exported {alert_count} alerts")
        
        print("📊 Exporting Relationships...")
        hd_count, drug_link_count, pharm_count, dp_count = export_relationships(wb)
        print(f"   ✓ Exported {hd_count} hospital-doctor links")
        print(f"   ✓ Exported {drug_link_count} hospital-drug links")
        print(f"   ✓ Exported {pharm_count} hospital-pharmacy links")
        print(f"   ✓ Exported {dp_count} doctor-patient links")
        
        # Create summary
        stats = {
            "Users": user_count,
            "Drugs": drug_count,
            "Patients": patient_count,
            "Alerts": alert_count,
            "Hospital-Doctor Links": hd_count,
            "Hospital-Drug Links": drug_link_count,
            "Hospital-Pharmacy Links": pharm_count,
            "Doctor-Patient Links": dp_count
        }
        export_summary(wb, stats)
        
        # Save
        filename = f"InteLeYzer_Database_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        print(f"\n✅ Excel file created: {filename}")
        print("="*60 + "\n")

if __name__ == "__main__":
    main()
