"""
Pharmacy Reports Routes
Handles submission, validation, and management of pharmacy safety reports
"""
from flask import Blueprint, request, jsonify, session
from datetime import datetime
from pv_backend.models import db
from pv_backend.models.pharmacy_report import (
    PharmacyReport, AnonymousReport, IdentifiedReport, AggregatedReport,
    ReportType, ReactionSeverity, ReactionOutcome, AgeGroup
)
from pv_backend.services.audit_service import log_action

pharmacy_report_bp = Blueprint('pharmacy_reports', __name__, url_prefix='/api/pharmacy/reports')

# ============================================================================
# SCHEMA VALIDATION
# ============================================================================

REPORT_SCHEMAS = {
    'anonymous': [
        'drug_name', 'batch_lot_number', 'dosage_form', 'date_of_dispensing',
        'reaction_category', 'severity', 'reaction_outcome', 'age_group', 'gender', 'additional_notes'
    ],
    'identified': [
        'drug_name', 'batch_lot_number', 'dosage_form', 'date_of_dispensing',
        'reaction_category', 'severity', 'reaction_outcome', 'age_group', 'gender',
        'internal_case_id', 'treating_hospital_reference', 'treating_doctor_name',
        'consent_verified', 'consent_date', 'additional_notes'
    ],
    'aggregated': [
        'drug_name', 'total_dispensed', 'total_reactions_reported',
        'mild_count', 'moderate_count', 'severe_count',
        'reporting_period_start', 'reporting_period_end', 'analysis_notes'
    ]
}

# ============================================================================
# SUBMISSION ENDPOINTS
# ============================================================================

@pharmacy_report_bp.route('/submit', methods=['POST'])
def submit_report():
    """
    Submit pharmacy safety report(s)
    Supports: Manual entry, Excel upload
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'message': 'No data provided'}), 400
        
        report_type = data.get('report_type', 'anonymous')
        entry_mode = data.get('entry_mode', 'manual')
        records = data.get('records', [])
        
        # Validate report type
        if report_type not in REPORT_SCHEMAS:
            return jsonify({'success': False, 'message': f'Invalid report type: {report_type}'}), 400
        
        if not records:
            return jsonify({'success': False, 'message': 'No records provided'}), 400
        
        # Validate schema for each record
        schema_columns = REPORT_SCHEMAS[report_type]
        for idx, record in enumerate(records):
            missing_required = []
            for col in schema_columns:
                if col not in record or record[col] == '':
                    # Check if it's a required field
                    if col in ['drug_name', 'dosage_form', 'date_of_dispensing', 'reaction_category', 'severity', 'age_group']:
                        missing_required.append(col)
            
            if missing_required:
                return jsonify({
                    'success': False,
                    'message': f'Record {idx + 1}: Missing required fields: {", ".join(missing_required)}'
                }), 400
        
        # Get pharmacy ID from session
        pharmacy_id = session.get('user_id')
        if not pharmacy_id:
            return jsonify({'success': False, 'message': 'Not authenticated'}), 401
        
        # Create submission records
        submission_id = f"SUB-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"
        created_records = []
        
        for record in records:
            try:
                if report_type == 'anonymous':
                    report = AnonymousReport(
                        report_type=ReportType.ANONYMOUS,
                        pharmacy_id=pharmacy_id,
                        drug_name=record.get('drug_name'),
                        drug_batch_number=record.get('batch_lot_number'),
                        reaction_description=record.get('reaction_category'),
                        reaction_severity=ReactionSeverity(record.get('severity', 'mild')),
                        reaction_outcome=ReactionOutcome(record.get('reaction_outcome', 'unknown')) if record.get('reaction_outcome') else None,
                        age_group=AgeGroup(record.get('age_group', 'unknown')),
                        gender=record.get('gender')
                    )
                
                elif report_type == 'identified':
                    report = IdentifiedReport(
                        report_type=ReportType.IDENTIFIED,
                        pharmacy_id=pharmacy_id,
                        drug_name=record.get('drug_name'),
                        drug_batch_number=record.get('batch_lot_number'),
                        reaction_description=record.get('reaction_category'),
                        reaction_severity=ReactionSeverity(record.get('severity', 'mild')),
                        reaction_outcome=ReactionOutcome(record.get('reaction_outcome', 'unknown')) if record.get('reaction_outcome') else None,
                        age_group=AgeGroup(record.get('age_group', 'unknown')),
                        gender=record.get('gender'),
                        internal_case_id=record.get('internal_case_id'),
                        treating_hospital_reference=record.get('treating_hospital_reference'),
                        treating_doctor_name=record.get('treating_doctor_name'),
                        follow_up_required=False
                    )
                
                elif report_type == 'aggregated':
                    report = AggregatedReport(
                        report_type=ReportType.AGGREGATED,
                        pharmacy_id=pharmacy_id,
                        drug_name=record.get('drug_name'),
                        reaction_description=f"Aggregated report for {record.get('drug_name')}",
                        report_count=record.get('total_reactions_reported', 1),
                        severity_distribution={
                            'mild': record.get('mild_count', 0),
                            'moderate': record.get('moderate_count', 0),
                            'severe': record.get('severe_count', 0)
                        }
                    )
                
                db.session.add(report)
                created_records.append(report)
            
            except Exception as e:
                db.session.rollback()
                return jsonify({
                    'success': False,
                    'message': f'Error processing record: {str(e)}'
                }), 500
        
        # Commit all records
        try:
            db.session.commit()
            
            # Log submission
            log_action(
                user_id=pharmacy_id,
                action='REPORT_SUBMITTED',
                details={
                    'submission_id': submission_id,
                    'report_type': report_type,
                    'entry_mode': entry_mode,
                    'record_count': len(created_records)
                }
            )
            
            return jsonify({
                'success': True,
                'submission_id': submission_id,
                'message': f'Successfully submitted {len(created_records)} records',
                'record_count': len(created_records)
            }), 201
        
        except Exception as e:
            db.session.rollback()
            return jsonify({
                'success': False,
                'message': f'Database error: {str(e)}'
            }), 500
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Server error: {str(e)}'
        }), 500


@pharmacy_report_bp.route('/validate-excel', methods=['POST'])
def validate_excel():
    """
    Validate Excel file schema before submission
    """
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['file']
        report_type = request.form.get('report_type', 'anonymous')
        
        if report_type not in REPORT_SCHEMAS:
            return jsonify({'success': False, 'error': f'Invalid report type: {report_type}'}), 400
        
        # Read Excel file
        import openpyxl
        try:
            workbook = openpyxl.load_workbook(file)
            worksheet = workbook.active
            
            # Get headers from first row
            headers = []
            for cell in worksheet[1]:
                if cell.value:
                    headers.append(cell.value)
            
            # Check if all required columns exist
            required_columns = REPORT_SCHEMAS[report_type]
            missing_columns = [col for col in required_columns if col not in headers]
            
            if missing_columns:
                return jsonify({
                    'success': False,
                    'error': f'Missing required columns: {", ".join(missing_columns)}',
                    'required_columns': required_columns
                }), 400
            
            # Read data rows
            preview_rows = []
            total_rows = 0
            
            for idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 1):
                if idx > 10:  # Preview first 10 rows
                    break
                
                row_data = {}
                for col_idx, header in enumerate(headers):
                    row_data[header] = row[col_idx] if col_idx < len(row) else None
                
                preview_rows.append(row_data)
                total_rows = idx
            
            # Count total rows
            total_rows = worksheet.max_row - 1  # Exclude header
            
            return jsonify({
                'success': True,
                'total_rows': total_rows,
                'preview_rows': preview_rows,
                'column_mapping': {h: h for h in headers}
            }), 200
        
        except Exception as e:
            return jsonify({
                'success': False,
                'error': f'Error reading Excel file: {str(e)}'
            }), 400
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Server error: {str(e)}'
        }), 500


# ============================================================================
# RETRIEVAL ENDPOINTS
# ============================================================================

@pharmacy_report_bp.route('/history', methods=['GET'])
def get_submission_history():
    """
    Get submission history for current pharmacy
    """
    try:
        pharmacy_id = session.get('user_id')
        if not pharmacy_id:
            return jsonify({'success': False, 'message': 'Not authenticated'}), 401
        
        limit = request.args.get('limit', 50, type=int)
        offset = request.args.get('offset', 0, type=int)
        
        # Query reports
        reports = PharmacyReport.query.filter_by(pharmacy_id=pharmacy_id)\
            .order_by(PharmacyReport.created_at.desc())\
            .limit(limit)\
            .offset(offset)\
            .all()
        
        return jsonify({
            'success': True,
            'reports': [r.to_dict() for r in reports],
            'total': len(reports)
        }), 200
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        }), 500


@pharmacy_report_bp.route('/<int:report_id>', methods=['GET'])
def get_report_detail(report_id):
    """
    Get detailed information about a specific report
    """
    try:
        pharmacy_id = session.get('user_id')
        if not pharmacy_id:
            return jsonify({'success': False, 'message': 'Not authenticated'}), 401
        
        report = PharmacyReport.query.filter_by(id=report_id, pharmacy_id=pharmacy_id).first()
        
        if not report:
            return jsonify({'success': False, 'message': 'Report not found'}), 404
        
        return jsonify({
            'success': True,
            'report': report.to_dict()
        }), 200
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        }), 500


# ============================================================================
# COMPLIANCE SCORING
# ============================================================================

@pharmacy_report_bp.route('/compliance-score', methods=['GET'])
def get_compliance_score():
    """
    Get compliance score for current pharmacy
    Scores based on:
    - On-time submissions
    - Alert acknowledgments
    - Data quality
    - Schema compliance
    """
    try:
        pharmacy_id = session.get('user_id')
        if not pharmacy_id:
            return jsonify({'success': False, 'message': 'Not authenticated'}), 401
        
        # Calculate compliance score
        score = 100  # Start with perfect score
        
        # Count recent submissions
        recent_reports = PharmacyReport.query.filter_by(pharmacy_id=pharmacy_id)\
            .filter(PharmacyReport.created_at >= datetime.utcnow().replace(day=1))\
            .count()
        
        # Deduct for no submissions this month
        if recent_reports == 0:
            score -= 10
        
        # Get compliance status
        if score >= 80:
            status = 'Compliant'
            status_color = 'green'
        elif score >= 60:
            status = 'Attention Required'
            status_color = 'yellow'
        else:
            status = 'Non-compliant'
            status_color = 'red'
        
        return jsonify({
            'success': True,
            'compliance_score': score,
            'status': status,
            'status_color': status_color,
            'last_updated': datetime.utcnow().isoformat()
        }), 200
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        }), 500
